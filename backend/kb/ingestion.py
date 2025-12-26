from typing import List, Dict, Any, Optional
import os
import re
import html
from pypdf import PdfReader
from .splitters import (
    NormalSplitter,
    AdaptiveSplitter,
)
from .splitters.splitter_table import TableSplitter
from .knowledge_base import FileInfo


def _table_to_markdown(rows: List[List[Any]]) -> str:
    """将二维表格行列数据转换为 Markdown 表格字符串"""
    if not rows:
        return ""

    normalized: List[List[str]] = []
    max_cols = 0
    for r in rows:
        if r is None:
            continue
        cols = []
        for c in (r or []):
            s = "" if c is None else str(c)
            s = re.sub(r"\s+", " ", s).strip()
            cols.append(s)
        max_cols = max(max_cols, len(cols))
        normalized.append(cols)

    if max_cols <= 0:
        return ""

    def _pad(row: List[str]) -> List[str]:
        if len(row) >= max_cols:
            return row[:max_cols]
        return row + [""] * (max_cols - len(row))

    normalized = [_pad(r) for r in normalized if r]
    if not normalized:
        return ""

    header = normalized[0]
    body = normalized[1:] if len(normalized) > 1 else []
    sep = ["---"] * max_cols

    out_lines = []
    out_lines.append("| " + " | ".join(header) + " |")
    out_lines.append("| " + " | ".join(sep) + " |")
    for r in body:
        out_lines.append("| " + " | ".join(r) + " |")
    return "\n".join(out_lines).strip()


def _extract_pdf_tables(pdf_path: str) -> Dict[int, List[str]]:
    """尝试从 PDF 中提取表格并返回按页分组的 Markdown 表格列表（无依赖则返回空）"""
    try:
        import pdfplumber  # type: ignore
    except Exception:
        return {}

    tables_by_page: Dict[int, List[str]] = {}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    continue
                rendered: List[str] = []
                for t in tables:
                    md = _table_to_markdown(t or [])
                    if md:
                        rendered.append(md)
                if rendered:
                    tables_by_page[page_idx] = rendered
    except Exception:
        return {}

    return tables_by_page


def read_pdf_text(pdf_path: str, include_tables: bool = True) -> str:
    """读取PDF文件的全部文本内容并返回字符串

    - 参数 `pdf_path`：PDF文件的绝对或相对路径
    - 返回：合并后的纯文本
    """
    reader = PdfReader(pdf_path)
    tables_by_page = _extract_pdf_tables(pdf_path) if include_tables else {}
    texts: List[str] = []
    for page_idx, page in enumerate(reader.pages):
        page_text = page.extract_text() or ""
        if include_tables:
            tables = tables_by_page.get(page_idx) or []
            if tables:
                blocks = []
                for i, md in enumerate(tables, start=1):
                    blocks.append(f"[Table][Page {page_idx + 1}][#{i}]\n{md}")
                page_text = (page_text + "\n\n" + "\n\n".join(blocks)).strip()
        texts.append(page_text)
    return "\n".join(texts).strip()

    
def read_chm_text(chm_path: str) -> str:
    """读取 CHM 文件为纯文本（使用 pychm + beautifulsoup4；无系统命令回退）。

    - 依赖：`pychm`（解析 CHM），`beautifulsoup4`（解析 HTML）
    - 若未安装依赖，会抛出明确的错误提示以指导安装
    - 文本解析：去除脚本/样式/注释，提取可读文本，适合后续拆分与索引
    """
    if not os.path.isfile(chm_path):
        raise FileNotFoundError(f"CHM 文件不存在：{chm_path}")

    try:
        from pychm import ChmFile  # type: ignore
    except Exception:
        raise RuntimeError(
            "缺少依赖：请安装 pychm 和 beautifulsoup4。\n"
            "pip install pychm beautifulsoup4"
        )

    chm = ChmFile(chm_path)
    try:
        content = chm.get_content()
    except Exception as e:
        raise RuntimeError(f"读取 CHM 内容失败: {e}")

    if not content:
        return ""

    if isinstance(content, (bytes, bytearray)):
        try:
            html_str = content.decode("utf-8")
        except Exception:
            html_str = content.decode("latin-1", errors="replace")
    else:
        html_str = str(content)

    try:
        from bs4 import BeautifulSoup  # type: ignore
        soup = BeautifulSoup(html_str, "html.parser")
        text = soup.get_text(" ", strip=True)
        return text
    except Exception:
        # bs4 不可用时，回退为简单的正则去标签（仍不使用系统命令）
        text = re.sub(r"(?is)<script.*?>.*?</script>", "", html_str)
        text = re.sub(r"(?is)<style.*?>.*?</style>", "", text)
        text = re.sub(r"(?is)<!--.*?-->", "", text)
        text = re.sub(r"(?is)<[^>]+>", " ", text)
        text = html.unescape(text)
        text = re.sub(r"\s+", " ", text).strip()
        return text


def read_excel_text(excel_path: str, max_rows_per_sheet: int = 2000, max_cols: int = 50) -> str:
    """读取 Excel（.xlsx/.xlsm）文件为纯文本，适合后续拆分与索引

    - 参数 `excel_path`：Excel 文件路径
    - 参数 `max_rows_per_sheet`：每个工作表最多保留的非空行数，避免超大表格导致内存/时延问题
    - 参数 `max_cols`：每行最多读取的列数，避免极宽表格影响可读性
    - 返回：按工作表分隔的文本，其中表格以 Markdown 表格形式输出
    """
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        raise RuntimeError(
            "缺少依赖：请安装 openpyxl 以读取 Excel 文件。\n"
            "pip install openpyxl"
        )

    def _cell_to_str(v: Any) -> str:
        s = "" if v is None else str(v)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    parts: List[str] = []
    for sheet in wb.worksheets:
        rows: List[List[Any]] = []
        kept_rows = 0
        try:
            for row in sheet.iter_rows(values_only=True):
                if kept_rows >= max_rows_per_sheet:
                    break
                cols = list(row[:max_cols]) if row else []
                if not any((c is not None and str(c).strip() != "") for c in cols):
                    continue
                kept_rows += 1
                rows.append([_cell_to_str(c) for c in cols])
        except Exception:
            continue

        if not rows:
            continue

        md = _table_to_markdown(rows)
        if not md:
            continue
        parts.append(f"[Sheet] {sheet.title}\n{md}")

    return "\n\n".join(parts).strip()


## 兼容保留：拆分函数已迁移至 kb.splitters 模块


def ingest_pdf(kb_controller, kb_id: int, pdf_path: str, chunk_size: int = 500, overlap: int = 100, use_llm_headings: Optional[bool] = None):
    """解析 PDF 并更新已存在文件的片段信息，不再创建文件记录

    - `kb_controller`：持久化知识库控制器实例
    - `kb_id`：知识库ID
    - `pdf_path`：PDF文件路径
    - `chunk_size` 与 `overlap`：回退分割参数
    - 返回：更新后的文件元信息对象（FileInfo）
    """
    text = read_pdf_text(pdf_path)
    use_llm = (
        bool(str(os.getenv("INGEST_USE_LLM_HEADING", "")).lower() in {"1", "true", "yes"})
        if use_llm_headings is None else bool(use_llm_headings)
    )
    adaptive_chunks = AdaptiveSplitter(use_llm=use_llm).split(text)
    if adaptive_chunks and adaptive_chunks[0].get("metadata", {}).get("number") == "" and adaptive_chunks[0].get("metadata", {}).get("type") == "toc":
        chunks = adaptive_chunks
    else:
        if adaptive_chunks and adaptive_chunks[0].get("metadata", {}).get("number"):
            chunks = adaptive_chunks
        else:
            chunks = NormalSplitter(chunk_size=chunk_size, overlap=overlap).split(text)
    filename = pdf_path.split("/")[-1].split("\\")[-1]
    meta = kb_controller._load_files(kb_id)
    files = meta.get("files", [])
    record = None
    for f in files:
        if str(f.get("filename")) == filename:
            record = f
            break
    if record is None:
        raise RuntimeError(f"文件未在知识库中登记：{filename}")
    file_id = int(record.get("id"))
    record["chunk_count"] = len(chunks)
    record["status"] = "done"
    kb_controller._save_files(kb_id, meta)
    kb_controller.save_chunks(kb_id, file_id=file_id, chunks=chunks)
    return FileInfo(id=file_id, filename=filename, chunk_count=len(chunks), status="done")


def ingest_excel(
    kb_controller,
    kb_id: int,
    excel_path: str,
    max_rows_per_sheet: int = 5000,
    max_cols: int = 50,
    use_llm_summary: Optional[bool] = None,
    max_rows_per_chunk: int = 80,
    max_chars_per_chunk: int = 8000,
):
    """解析 Excel 并更新已存在文件的片段信息，不再创建文件记录

    - 拆分结构：表格名称（Excel 文件名） + Sheet 名称 + Sheet 内容（Markdown 表格）
    - 可选：基于“表格名称 + Sheet 名称 + 表头字段”调用 LLM 生成摘要并放在片段开头
    - 返回：更新后的文件元信息对象（FileInfo）
    """
    text = read_excel_text(
        excel_path,
        max_rows_per_sheet=max_rows_per_sheet,
        max_cols=max_cols,
    )
    use_llm = (
        bool(str(os.getenv("INGEST_USE_LLM_TABLE_SUMMARY", "")).lower() in {"1", "true", "yes"})
        if use_llm_summary is None else bool(use_llm_summary)
    )
    filename = excel_path.split("/")[-1].split("\\")[-1]
    table_name = os.path.splitext(filename)[0]

    chunks = TableSplitter(
        table_name=table_name,
        use_llm_summary=use_llm,
        max_rows_per_chunk=max_rows_per_chunk,
        max_chars_per_chunk=max_chars_per_chunk,
    ).split(text)
    if not chunks:
        chunks = [{
            "content": f"[Table] {table_name}\n[ExcelEmpty] 未读取到任何非空表格数据",
            "metadata": {"type": "table", "table_name": table_name, "sheet_name": "", "part_index": 1, "part_count": 1, "header": []},
        }]

    meta = kb_controller._load_files(kb_id)
    files = meta.get("files", [])
    record = None
    for f in files:
        if str(f.get("filename")) == filename:
            record = f
            break
    if record is None:
        raise RuntimeError(f"文件未在知识库中登记：{filename}")
    file_id = int(record.get("id"))
    record["chunk_count"] = len(chunks)
    record["status"] = "done"
    kb_controller._save_files(kb_id, meta)
    kb_controller.save_chunks(kb_id, file_id=file_id, chunks=chunks)
    return FileInfo(id=file_id, filename=filename, chunk_count=len(chunks), status="done")

