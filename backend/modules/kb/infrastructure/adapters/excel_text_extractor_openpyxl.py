from __future__ import annotations

import re
from typing import Any, List

from backend.modules.kb.domain.ports import ExcelTextExtractorPort


class OpenpyxlExcelTextExtractor(ExcelTextExtractorPort):
    def extract_text(self, excel_path: str, *, max_rows_per_sheet: int, max_cols: int) -> str:
        import os

        if not os.path.isfile(excel_path):
            raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:
            raise RuntimeError("缺少依赖：请安装 openpyxl 以读取 Excel 文件") from e

        def _cell_to_str(v: Any) -> str:
            s = "" if v is None else str(v)
            s = re.sub(r"\s+", " ", s).strip()
            return s

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        parts: List[str] = []
        for sheet in wb.worksheets:
            rows: List[List[Any]] = []
            kept_rows = 0
            try:
                for row in sheet.iter_rows(values_only=True):
                    if kept_rows >= int(max_rows_per_sheet):
                        break
                    cols = list(row[: int(max_cols)]) if row else []
                    if not any((c is not None and str(c).strip() != "") for c in cols):
                        continue
                    kept_rows += 1
                    rows.append([_cell_to_str(c) for c in cols])
            except Exception:
                continue

            if not rows:
                continue
            md = self._table_to_markdown(rows)
            if not md:
                continue
            parts.append(f"[Sheet] {sheet.title}\n{md}")
        return "\n\n".join(parts).strip()

    def _table_to_markdown(self, rows: List[List[str]]) -> str:
        if not rows:
            return ""
        width = max(len(r) for r in rows)
        if width <= 0:
            return ""
        norm: List[List[str]] = []
        for r in rows:
            rr = [(c or "").strip() for c in (r or [])]
            if len(rr) < width:
                rr = rr + [""] * (width - len(rr))
            norm.append(rr[:width])
        header = norm[0]
        sep = ["---"] * width
        body = norm[1:] if len(norm) > 1 else []
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(sep) + " |",
        ]
        for r in body:
            lines.append("| " + " | ".join(r) + " |")
        return "\n".join(lines).strip()

