from typing import Any, Dict, List, Optional
import html
import logging
import os
import re
import time
from urllib.parse import unquote, urlparse

from .splitters import AdaptiveSplitter, NormalSplitter
from .splitters.splitter_table import TableSplitter
from backend.modules.kb.domain.models import ChunkMetadata, FileInfo, KnowledgeChunk
from .services.image_caption import caption_jobs

logger = logging.getLogger(__name__)




def read_pdf_markdown_with_images(pdf_path: str, image_dir: str) -> str:
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF 文件不存在：{pdf_path}")

    try:
        import pymupdf4llm  # type: ignore
    except Exception as e:
        raise RuntimeError("缺少依赖：请安装 pymupdf 与 pymupdf4llm") from e

    os.makedirs(image_dir, exist_ok=True)
    md = pymupdf4llm.to_markdown(
        pdf_path,
        write_images=True,
        image_path=image_dir,
        use_ocr=False,
    )
    if isinstance(md, list):
        page_texts: List[str] = []
        for page in md:
            page_texts.append(str(page.get("text")).strip())
        return "\n\n".join(t for t in page_texts if t).strip()
    return str(md).strip()

def _table_to_markdown(rows: List[List[str]]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    if width <= 0:
        return ""
    norm: List[List[str]] = []
    for r in rows:
        rr = [(c or "").strip() for c in (r or [])]
        if len(rr) < width:
            rr = rr + [""] * (width - len(rr))
        norm.append(rr[:width])
    header = norm[0]
    sep = ["---"] * width
    body = norm[1:] if len(norm) > 1 else []
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines).strip()


def read_excel_text(excel_path: str, max_rows_per_sheet: int = 2000, max_cols: int = 50) -> str:
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("缺少依赖：请安装 openpyxl 以读取 Excel 文件") from e

    def _cell_to_str(v: Any) -> str:
        s = "" if v is None else str(v)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    parts: List[str] = []
    for sheet in wb.worksheets:
        rows: List[List[Any]] = []
        kept_rows = 0
        try:
            for row in sheet.iter_rows(values_only=True):
                if kept_rows >= max_rows_per_sheet:
                    break
                cols = list(row[:max_cols]) if row else []
                if not any((c is not None and str(c).strip() != "") for c in cols):
                    continue
                kept_rows += 1
                rows.append([_cell_to_str(c) for c in cols])
        except Exception:
            continue

        if not rows:
            continue

        md = _table_to_markdown(rows)
        if not md:
            continue
        parts.append(f"[Sheet] {sheet.title}\n{md}")

    return "\n\n".join(parts).strip()


def read_chm_text(chm_path: str) -> str:
    if not os.path.isfile(chm_path):
        raise FileNotFoundError(f"CHM 文件不存在：{chm_path}")

    try:
        from pychm import ChmFile  # type: ignore
    except Exception as e:
        raise RuntimeError("缺少依赖：请安装 pychm 和 beautifulsoup4") from e

    chm = ChmFile(chm_path)
    try:
        content = chm.get_content()
    except Exception as e:
        raise RuntimeError(f"读取 CHM 内容失败: {e}")

    if not content:
        return ""

    if isinstance(content, (bytes, bytearray)):
        try:
            html_str = content.decode("utf-8")
        except Exception:
            html_str = content.decode("latin-1", errors="replace")
    else:
        html_str = str(content)

    try:
        from bs4 import BeautifulSoup  # type: ignore

        soup = BeautifulSoup(html_str, "html.parser")
        return soup.get_text(" ", strip=True)
    except Exception:
        text = re.sub(r"(?is)<script.*?>.*?</script>", "", html_str)
        text = re.sub(r"(?is)<style.*?>.*?</style>", "", text)
        text = re.sub(r"(?is)<!--.*?-->", "", text)
        text = re.sub(r"(?is)<[^>]+>", " ", text)
        text = html.unescape(text)
        text = re.sub(r"\s+", " ", text).strip()
        return text


def ingest_pdf(
    kb_controller,
    kb_id: int,
    pdf_path: str,
    chunk_size: int = 500,
    overlap: int = 100,
    use_llm_headings: Optional[bool] = None,
):
    from backend.modules.kb.application.services.file_ingestion import FileIngestionService, PdfIngestionOptions
    from backend.modules.kb.infrastructure.adapters.excel_text_extractor_openpyxl import OpenpyxlExcelTextExtractor
    from backend.modules.kb.infrastructure.adapters.image_captioner_legacy import LegacyImageCaptioner
    from backend.modules.kb.infrastructure.adapters.pdf_markdown_extractor_pymupdf4llm import PyMuPdf4LlmPdfMarkdownExtractor
    from backend.modules.kb.infrastructure.adapters.table_chunker_legacy import LegacyTableChunker

    ingestion = FileIngestionService(
        file_resolver=kb_controller,
        asset_paths=kb_controller,
        chunk_writer=kb_controller,
        pdf_extractor=PyMuPdf4LlmPdfMarkdownExtractor(),
        image_captioner=LegacyImageCaptioner(),
        excel_text_extractor=OpenpyxlExcelTextExtractor(),
        table_chunker=LegacyTableChunker(),
    )
    return ingestion.ingest_pdf(
        kb_id=int(kb_id),
        pdf_path=pdf_path,
        options=PdfIngestionOptions(chunk_size=int(chunk_size), overlap=int(overlap), use_llm_headings=use_llm_headings),
    )


def ingest_excel(
    kb_controller,
    kb_id: int,
    excel_path: str,
    max_rows_per_sheet: int = 5000,
    max_cols: int = 50,
    use_llm_summary: Optional[bool] = None,
    max_rows_per_chunk: int = 80,
    max_chars_per_chunk: int = 8000,
):
    from backend.modules.kb.application.services.file_ingestion import ExcelIngestionOptions, FileIngestionService
    from backend.modules.kb.infrastructure.adapters.excel_text_extractor_openpyxl import OpenpyxlExcelTextExtractor
    from backend.modules.kb.infrastructure.adapters.image_captioner_legacy import LegacyImageCaptioner
    from backend.modules.kb.infrastructure.adapters.pdf_markdown_extractor_pymupdf4llm import PyMuPdf4LlmPdfMarkdownExtractor
    from backend.modules.kb.infrastructure.adapters.table_chunker_legacy import LegacyTableChunker

    ingestion = FileIngestionService(
        file_resolver=kb_controller,
        asset_paths=kb_controller,
        chunk_writer=kb_controller,
        pdf_extractor=PyMuPdf4LlmPdfMarkdownExtractor(),
        image_captioner=LegacyImageCaptioner(),
        excel_text_extractor=OpenpyxlExcelTextExtractor(),
        table_chunker=LegacyTableChunker(),
    )
    return ingestion.ingest_excel(
        kb_id=int(kb_id),
        excel_path=excel_path,
        options=ExcelIngestionOptions(
            max_rows_per_sheet=int(max_rows_per_sheet),
            max_cols=int(max_cols),
            use_llm_summary=use_llm_summary,
            max_rows_per_chunk=int(max_rows_per_chunk),
            max_chars_per_chunk=int(max_chars_per_chunk),
        ),
    )
